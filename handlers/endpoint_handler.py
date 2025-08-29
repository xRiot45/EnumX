import csv
import json
import openpyxl
from openpyxl.styles import Alignment, Font

from utils.logger import Logger

logger = Logger()


class EndpointHandler:
    @staticmethod
    def save(results, filepath, format_type="json"):
        endpoints = results.get("endpoints", [])

        # --- All formats ---
        if format_type == "all":
            base = filepath.rsplit(".", 1)[0]
            for fmt in ["json", "csv", "txt", "xlsx", "html", "md"]:
                EndpointHandler.save(results, f"{base}.{fmt}", fmt)
            return

        # --- JSON format ---
        elif format_type == "json":
            with open(filepath, "w") as f:
                json.dump(results, f, indent=4)
            logger.info(f"Results successfully saved to {filepath} ({format_type.upper()})")

        # --- CSV format ---
        elif format_type == "csv":
            with open(filepath, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["URL", "Status", "Length", "Methods"])
                for ep in endpoints:
                    writer.writerow([ep["url"], ep["status"], ep["length"], ",".join(ep["methods"])])
            logger.info(f"Results successfully saved to {filepath} ({format_type.upper()})")

        # --- TXT format ---
        elif format_type == "txt":
            with open(filepath, "w") as f:
                for ep in endpoints:
                    f.write(f"{ep['url']} [{ep['status']}] ({ep['length']} bytes) Methods: {','.join(ep['methods'])}\n")
            logger.info(f"Results successfully saved to {filepath} ({format_type.upper()})")

        # --- XLSX format ---
        elif format_type == "xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Endpoint Results"

            header = ["URL", "Status", "Length", "Methods"]
            ws.append(header)

            for col in range(1, len(header) + 1):
                ws.cell(row=1, column=col).font = Font(bold=True)
                ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

            for ep in endpoints:
                ws.append([ep["url"], ep["status"], ep["length"], ",".join(ep["methods"])])

            wb.save(filepath)
            logger.info(f"Results successfully saved to {filepath} ({format_type.upper()})")

        # --- HTML format ---
        elif format_type == "html":
            with open(filepath, "w") as f:
                f.write("<!DOCTYPE html><html><head><meta charset='UTF-8'>\n")
                f.write(
                    "<style>table{border-collapse:collapse;width:100%;}"
                    "th,td{border:1px solid #ddd;padding:8px;}"
                    "th{background:#333;color:#fff;}</style></head><body>\n"
                )
                f.write("<h2>Endpoint Enumeration Results</h2>\n")
                f.write("<table>\n")
                f.write("<tr><th>URL</th><th>Status</th><th>Length</th><th>Methods</th></tr>\n")

                for ep in endpoints:
                    f.write("<tr>")
                    f.write(f"<td>{ep['url']}</td>")
                    f.write(f"<td>{ep['status']}</td>")
                    f.write(f"<td>{ep['length']}</td>")
                    f.write(f"<td>{','.join(ep['methods'])}</td>")
                    f.write("</tr>\n")

                f.write("</table></body></html>\n")
            logger.info(f"Results successfully saved to {filepath} ({format_type.upper()})")

        # --- Markdown format ---
        elif format_type == "md":
            with open(filepath, "w") as f:
                f.write("# Endpoint Enumeration Results\n\n")
                f.write("| URL | Status | Length | Methods |\n")
                f.write("|-----|--------|--------|---------|\n")
                for ep in endpoints:
                    f.write(f"| {ep['url']} | {ep['status']} | {ep['length']} | {','.join(ep['methods'])} |\n")
            logger.info(f"Results successfully saved to {filepath} ({format_type.upper()})")

        else:
            logger.error(f"Unsupported format: {format_type}")

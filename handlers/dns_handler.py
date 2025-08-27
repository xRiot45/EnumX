import csv
import json

import openpyxl
from openpyxl.styles import Alignment, Font

from utils.logger import Logger

logger = Logger()


class DNSHandler:
    @staticmethod
    def save(results, filepath, format_type="json"):
        subdomains = results.get("subdomains", [])

        # --- JSON format
        if format_type == "json":
            with open(filepath, "w") as f:
                json.dump(results, f, indent=4)
            logger.info(f"[DNS] Saved JSON → {filepath}")

        # --- CSV format ---
        elif format_type == "csv":
            with open(filepath, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Subdomain", "Record Type", "Class", "TTL", "Record Value"])
                for entry in subdomains:
                    sub = entry["subdomain"]
                    first_row = True
                    for rtype, values in entry.get("records", {}).items():
                        for val in values:
                            writer.writerow(
                                [
                                    sub if first_row else "",
                                    rtype,
                                    val.get("class", "IN"),
                                    val.get("ttl", ""),
                                    val.get("value", ""),
                                ]
                            )
                            first_row = False
            logger.info(f"[DNS] Saved CSV → {filepath}")

        # --- TXT format ---
        elif format_type == "txt":
            with open(filepath, "w") as f:
                for entry in subdomains:
                    f.write(f"{entry['subdomain']}\n")
                    for rtype, values in entry.get("records", {}).items():
                        for val in values:
                            f.write(f"  {rtype} {val.get('class','IN')} {val.get('ttl','')} → {val.get('value','')}\n")
            logger.info(f"[DNS] Saved TXT → {filepath}")

        # --- XLSX format ---
        elif format_type == "xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "DNS Results"

            header = ["Subdomain", "Record Type", "Class", "TTL", "Record Value"]
            ws.append(header)

            for col in range(1, len(header) + 1):
                ws.cell(row=1, column=col).font = Font(bold=True)
                ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

            row = 2
            for entry in subdomains:
                sub = entry["subdomain"]
                records = []
                for rtype, values in entry.get("records", {}).items():
                    for val in values:
                        records.append([rtype, val.get("class", "IN"), val.get("ttl", ""), val.get("value", "")])

                if not records:
                    continue

                start_row = row
                for rec in records:
                    ws.append([sub, *rec])
                    row += 1

                if row - start_row > 1:
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
                    ws.cell(start_row, 1).alignment = Alignment(vertical="top")

            wb.save(filepath)
            logger.info(f"[DNS] Saved XLSX with merged cells → {filepath}")
        
        # --- HTML format ---
        elif format_type == "html":
            with open(filepath, "w") as f:
                f.write("<!DOCTYPE html><html><head><meta charset='UTF-8'>\n")
                f.write("<style>table{border-collapse:collapse;width:100%;}"
                        "th,td{border:1px solid #ddd;padding:8px;}"
                        "th{background:#333;color:#fff;}</style></head><body>\n")
                f.write("<h2>DNS Enumeration Results</h2>\n")
                f.write("<table>\n")
                f.write("<tr><th>Subdomain</th><th>Record Type</th><th>Class</th><th>TTL</th><th>Record Value</th></tr>\n")

                for entry in subdomains:
                    sub = entry["subdomain"]
                    first_row = True
                    for rtype, values in entry.get("records", {}).items():
                        for val in values:
                            f.write("<tr>")
                            f.write(f"<td>{sub if first_row else ''}</td>")
                            f.write(f"<td>{rtype}</td>")
                            f.write(f"<td>{val.get('class','IN')}</td>")
                            f.write(f"<td>{val.get('ttl','')}</td>")
                            f.write(f"<td>{val.get('value','')}</td>")
                            f.write("</tr>\n")
                            first_row = False
                f.write("</table></body></html>\n")
            logger.info(f"[DNS] Saved HTML → {filepath}")
        
        # --- Markdown format ---
        elif format_type == "md":
            with open(filepath, "w") as f:
                f.write("# DNS Enumeration Results\n\n")
                f.write("| Subdomain | Record Type | Class | TTL | Record Value |\n")
                f.write("|-----------|-------------|-------|-----|--------------|\n")
                for entry in subdomains:
                    sub = entry["subdomain"]
                    first_row = True
                    for rtype, values in entry.get("records", {}).items():
                        for val in values:
                            f.write(
                                f"| {sub if first_row else ''} | {rtype} | {val.get('class','IN')} | {val.get('ttl','')} | {val.get('value','')} |\n"
                            )
                            first_row = False
            logger.info(f"[DNS] Saved Markdown → {filepath}")

        else:
            logger.error(f"Unsupported format: {format_type}")
